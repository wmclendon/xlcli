"""main.py - Main entry point for the xcli tool"""

from pathlib import Path
import typer
from rich.console import Console
from rich.table import Table
from openpyxl import load_workbook


app = typer.Typer()

console = Console()


@app.command(name="print")
def print_xlsx(
    file: str,
    sheet: str
    | None = typer.Option(None, help="Sheet name to read. Defaults to the first sheet"),
):
    """Read an Excel file and display its contents"""

    file_path = Path(file)
    if not file_path.exists():
        console.print(f"[bold red]File '{file}' not found.")
        raise typer.Exit(code=1)
    wb = load_workbook(file)
    if sheet:
        ws = wb[sheet]
    else:
        ws = wb.active

    table = Table(title=f"{file_path.stem} - {ws.title}", show_lines=True)
    for col in ws.iter_cols(1, ws.max_column):
        table.add_column(str(col[0].value))
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        table.add_row(*[str(cell) for cell in row])

    console.print(table)


@app.command()
def info(file: str):
    """Display information about an Excel file"""
    file_path = Path(file)
    if not file_path.exists():
        console.print(f"[bold red]File '{file}' not found.")
        raise typer.Exit(code=1)
    wb = load_workbook(file)
    table = Table(title=f"{file_path.stem} - Info", show_lines=True)
    table.add_column("Property")
    table.add_column("Value")
    table.add_row("Sheets", str(len(wb.sheetnames)))
    table.add_row("Default Sheet", wb.active.title)
    table.add_row("Sheet Names", "\n".join(wb.sheetnames))
    console.print(table)


if __name__ == "__main__":
    app()
